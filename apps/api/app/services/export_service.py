from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session

from app.repositories.category_repository import CategoryRepository
from app.repositories.product_repository import ProductRepository

##* aca decidimos como se veran los elementos visuales 
##* headers es la lista con los nombres de las columnas en orden
HEADERS = [
    "Nombre", "SKU", "Categoría", "Cantidad", "Stock mínimo",
    "Precio unitario", "Valor total", "Ubicación", "Estado", "Alerta de stock bajo",
]

##* define el color de fondo para los encabezados
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True) ##* define como sera la letra, blanca y en negrita
LOW_STOCK_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") ##* definimos un color rojo de fondo para los que tengan poco stock

##* inicializamos ambos repositorios uno para productos y otro para categorias
class ExportService:
    def __init__(self, session: Session):
        self.product_repository = ProductRepository(session)
        self.category_repository = CategoryRepository(session)

    def build_products_workbook(self, user_id: int) -> BytesIO:
        products = self.product_repository.list_all(user_id)
        categories_by_id = {c.id: c.name for c in self.category_repository.list_all(user_id)}

        ##* creamos un archivo excel en blanco, seleccionamos la primera hoja y la llamamos INVENTARIO
        ##* luego inserta la primera fila con los titulos de las columnas sheet.append(HEADERS)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventario"

        sheet.append(HEADERS)
        for col_index in range(1, len(HEADERS) + 1):
            cell = sheet.cell(row=1, column=col_index)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center") ##* Recorre las celdas de la fila 1 (los encabezados) 
            ##*una por una para aplicarles el color de fondo índigo, la letra blanca en negrita y centrar el texto.



            ##! Cálculos rápidos: Verifica si la cantidad actual es menor o igual al stock mínimo (is_low_stock)
            ##!  y calcula el valor total multiplicando cantidad por precio.
        for product in products:
            is_low_stock = product.quantity <= product.min_stock
            total_value = product.quantity * product.price



            ##! Armar la fila: Creamos una lista (row) con las propiedades del producto ordenadas tal cual 
            ##!los encabezados. Si no tiene categoría, pone "Sin categoría". Si la ubicación es None, pone un texto vacío "".
            row = [
                product.name,
                product.sku,
                categories_by_id.get(product.category_id, "Sin categoría"),
                product.quantity,
                product.min_stock,
                float(product.price),
                float(total_value),
                product.location or "",
                product.status,
                "Sí" if is_low_stock else "No",
            ]
            sheet.append(row)



            ##! Si el producto tiene stock bajo, detecta en qué número de fila acaba de ser insertado (sheet.max_row) 
            ##! y recorre todas las celdas de esa fila en particular para pintarles el fondo de color rojo claro. Así, salta a la vista inmediatamente al abrir el Excel.
            if is_low_stock:
                row_index = sheet.max_row
                for col_index in range(1, len(HEADERS) + 1):
                    sheet.cell(row=row_index, column=col_index).fill = LOW_STOCK_FILL

        ##* formato de moneda en las columnas de precio y valor total
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=6).number_format = "$#,##0.00"
            sheet.cell(row=row_index, column=7).number_format = "$#,##0.00"

        ##* ancho de columnas ajustado al contenido, para que el Excel se vea prolijo al abrirlo
        for col_index, header in enumerate(HEADERS, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = max(len(header) + 4, 14)

        sheet.freeze_panes = "A2"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
